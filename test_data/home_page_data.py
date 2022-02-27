import openpyxl


class HomePageData:
    test_home_page_data = [
        {"firstname": "Rose", "email": "rose@gmail.com", "gender": "Female"},
        {"firstname": "Ash", "email": "ash@gmail.com", "gender": "Male"}
    ]

    test_home_page_data_tuple = [
        ('Rose', 'rose@gmail.com', 'Female'),
        ('Ash', 'ash@gmail.com', 'Male')
    ]

    @staticmethod
    def get_data_excel(test_case):
        book = openpyxl.load_workbook(
            'D:\\Dev\\Selenium\\udemy\\91 E2E Desktop Test\\test_data\\xlsx\\python-demo.xlsx'
        )
        sheet = book.active

        info_dict = {}

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case:
                for j in range(2, sheet.max_column + 1):
                    info_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [info_dict]
