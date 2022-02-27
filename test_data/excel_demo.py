import openpyxl


book = openpyxl.load_workbook('D:\\Dev\\Selenium\\udemy\\91 E2E Desktop Test\\test_data\\xlsx\\python-demo.xlsx')
sheet = book.active
cell = sheet.cell(row=1, column=2)

info_dict = {}

# print(cell.value)
#
# sheet.cell(row=2, column=2).value = 'Ross'
#
# print(sheet.cell(row=2, column=2).value)
#
# print(sheet.max_row)
# print(sheet.max_column)

# for i in range(1, sheet.max_row + 1):
#     for j in range(1, sheet.max_column + 1):
#         print(sheet.cell(row=i, column=j).value)

# for i in range(1, sheet.max_column + 1):
#     for j in range(1, sheet.max_row + 1):
#         print(sheet.cell(row=j, column=i).value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == 'test5':
        for j in range(2, sheet.max_column + 1):
            info_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(info_dict)